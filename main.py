import os
import re
import shutil
import openpyxl
import unicodedata
from datetime import datetime
from collections import defaultdict
from openpyxl.styles import PatternFill

# ==========================================
# 1. CONFIGURACIONES Y FUNCIONES COMPARTIDAS
# ==========================================

MESES_ES = {
    "01": "ENERO", "02": "FEBRERO", "03": "MARZO", "04": "ABRIL",
    "05": "MAYO", "06": "JUNIO", "07": "JULIO", "08": "AGOSTO",
    "09": "SEPTIEMBRE", "10": "OCTUBRE", "11": "NOVIEMBRE", "12": "DICIEMBRE",
}

def _convertir_num(valor):
    """Limpia y convierte textos de moneda a números matemáticos."""
    if valor is None or str(valor).strip() == "":
        return 0
    if isinstance(valor, (int, float)):
        return valor
    if isinstance(valor, str):
        try:
            clean_val = valor.replace("$", "").replace(".", "").replace(",", ".")
            return float(clean_val)
        except ValueError:
            return 0
    return 0

def clean_header(text):
    """Limpia encabezados para buscarlos sin importar espacios o mayúsculas."""
    return str(text).strip().upper().replace(" ", "")

def normalizar(texto):
    """Normalización estricta para cruzar métodos de pago."""
    if not texto:
        return ""
    texto = str(texto).upper().replace(" ", "")
    texto = texto.replace("DE", "")
    texto = ''.join(c for c in unicodedata.normalize('NFD', texto) if unicodedata.category(c) != 'Mn')
    texto = re.sub(r'[^A-Z0-9]', '', texto)
    return texto

def mover_procesados(archivos):
    """Mueve los archivos ya procesados a una carpeta para evitar duplicarlos."""
    carpeta_procesados = "PROCESADOS"
    if not os.path.exists(carpeta_procesados):
        os.makedirs(carpeta_procesados)
        
    for f_origen in archivos:
        ruta_destino = os.path.join(carpeta_procesados, f_origen)
        if os.path.exists(ruta_destino):
            os.remove(ruta_destino)
        shutil.move(f_origen, carpeta_procesados)
        print(f"  - Archivo {f_origen} movido a la carpeta /{carpeta_procesados}")

# ==========================================
# 2. MÓDULO: FACTURAS Y NOTAS DE CRÉDITO
# ==========================================

def procesar_facturas_y_nc():
    print("\n" + "="*50)
    print("▶ INICIANDO: TRASPASO DE FACTURAS Y NOTAS DE CRÉDITO")
    print("="*50)
    
    archivos_a_procesar = []
    archivos_originales = []
    
    patron_factura = re.compile(r"FACTURAS\s+(\d+)\s+(\d{2})-(\d{2})", re.IGNORECASE)
    patron_nc = re.compile(r"NC\s+TOTAL\s+(\d+)\s+(\d{2})-(\d{2})", re.IGNORECASE)

    for file in os.listdir():
        if file.lower().endswith((".xls", ".xlsx")) and not file.startswith("~"):
            match_fact = patron_factura.search(file)
            match_nc = patron_nc.search(file)
            
            if match_fact:
                local, dia, mes = match_fact.groups()
                mes_nombre = MESES_ES.get(mes)
                if mes_nombre:
                    archivos_a_procesar.append(('FACTURA', local, int(dia), mes_nombre, file))
                    archivos_originales.append(file)
            elif match_nc:
                local, dia, mes = match_nc.groups()
                mes_nombre = MESES_ES.get(mes)
                if mes_nombre:
                    archivos_a_procesar.append(('NC', local, int(dia), mes_nombre, file))
                    archivos_originales.append(file)

    if not archivos_a_procesar:
        print("⚠️ No se encontraron archivos de 'FACTURAS' ni 'NC TOTAL' nuevos.")
        return

    datos_por_dia = {}
    
    for tipo_doc, local, day, mes_nombre, fname in archivos_a_procesar:
        print(f"Leyendo [{tipo_doc}] desde: {fname}...")
        try:
            wb_origen = openpyxl.load_workbook(fname, data_only=True)
            ws_origen = wb_origen.active

            header_map = {}
            for col in range(1, ws_origen.max_column + 1):
                val = ws_origen.cell(row=1, column=col).value
                if val:
                    header_map[clean_header(val)] = col
            
            col_serie = header_map.get("SERIE") or header_map.get("FECHA")
            col_nombre = header_map.get("NOMBRE") or header_map.get("CLIENTE")
            col_numero = header_map.get("NÚMERO") or header_map.get("NUMERO") or header_map.get("FACTURA")
            col_total = header_map.get("TOTAL") or header_map.get("MONTO")

            if not all([col_serie, col_nombre, col_numero, col_total]):
                print(f"  ⚠️ Faltan columnas clave en {fname}. Saltando.")
                continue

            if day not in datos_por_dia:
                datos_por_dia[day] = []

            for row in range(2, ws_origen.max_row + 1):
                serie = str(ws_origen.cell(row=row, column=col_serie).value or "").strip()
                nombre = str(ws_origen.cell(row=row, column=col_nombre).value or "").strip()
                numero_val = ws_origen.cell(row=row, column=col_numero).value
                numero = str(numero_val).strip() if numero_val is not None else ""
                total = ws_origen.cell(row=row, column=col_total).value

                if not serie or serie.upper() == "NONE" or "TOTAL" in serie.upper():
                    continue

                datos_por_dia[day].append({
                    'tipo': tipo_doc,
                    'serie': serie,
                    'nombre': nombre,
                    'numero': numero,
                    'total': _convertir_num(total)
                })
        except Exception as e:
            print(f"  ❌ Error al leer {fname}: {e}")

    if not datos_por_dia:
        print("\n⚠️ No se extrajeron datos válidos para procesar.")
        return

    mes_nombre_dest = archivos_a_procesar[0][3]
    ano_actual = datetime.now().year
    archivo_destino = f"{mes_nombre_dest} {ano_actual}.xlsx"

    print(f"\nAbriendo archivo maestro: {archivo_destino}...")
    try:
        ventas_wb = openpyxl.load_workbook(archivo_destino)
        hoja_nombre = "FACTURA A COBRAR"
        ventas_ws = ventas_wb[hoja_nombre]
    except Exception as e:
        print(f"❌ Error al cargar maestro: {e}. Asegúrate de que esté CERRADO.")
        return

    col_dest_fecha = 1
    col_dest_nombre = 3
    col_dest_n_fact = 4
    col_dest_monto_f = 5
    col_dest_n_nc = 6
    col_dest_monto_nc = 7

    cont_facturas = 0
    cont_nc = 0

    for dia in sorted(datos_por_dia.keys(), reverse=True):
        datos_del_dia = datos_por_dia[dia]
        
        # --- AQUÍ ESTÁ LA MEJORA ---
        # Ordenamos la lista para que FACTURA quede antes que NC
        datos_del_dia.sort(key=lambda x: 0 if x['tipo'] == 'FACTURA' else 1)
        # ---------------------------

        fila_marcador = None
        col_marcador = None
        
        for r in range(1, ventas_ws.max_row + 1):
            encontrado = False
            for c in range(1, 4): 
                celda_val = ventas_ws.cell(row=r, column=c).value
                if celda_val is not None:
                    if isinstance(celda_val, (int, float)) and int(celda_val) == int(dia):
                        encontrado = True
                    elif isinstance(celda_val, str):
                        texto_limpio = celda_val.strip()
                        if texto_limpio == str(dia) or texto_limpio == f"{dia}.0":
                            encontrado = True
                if encontrado:
                    fila_marcador = r
                    col_marcador = c
                    break
            if encontrado:
                break
        
        if fila_marcador:
            fila_insercion = fila_marcador
            for fila_arriba in range(fila_marcador - 1, 0, -1):
                val_arriba = ventas_ws.cell(row=fila_arriba, column=col_marcador).value
                es_fecha = False
                
                if val_arriba is not None:
                    if isinstance(val_arriba, datetime):
                        es_fecha = True
                    elif isinstance(val_arriba, str):
                        txt = val_arriba.strip().upper()
                        if "FECHA" in txt or "SERIE" in txt or re.search(r'\d{1,2}[-/]\d{1,2}', txt):
                            es_fecha = True
                
                if es_fecha:
                    val_mas_arriba = ventas_ws.cell(row=fila_arriba - 1, column=col_marcador).value if fila_arriba > 1 else None
                    es_mas_arriba_fecha = False
                    if val_mas_arriba is not None:
                        if isinstance(val_mas_arriba, datetime): es_mas_arriba_fecha = True
                        elif isinstance(val_mas_arriba, str):
                            txt_mas = val_mas_arriba.strip().upper()
                            if "FECHA" in txt_mas or "SERIE" in txt_mas or re.search(r'\d{1,2}[-/]\d{1,2}', txt_mas):
                                es_mas_arriba_fecha = True
                    
                    if not es_mas_arriba_fecha:
                        fila_insercion = fila_arriba + 1 
                        break

            cantidad_a_insertar = len(datos_del_dia)
            ventas_ws.insert_rows(fila_insercion, amount=cantidad_a_insertar)
            
            for i, dato in enumerate(datos_del_dia):
                fila_escritura = fila_insercion + i
                
                ventas_ws.cell(row=fila_escritura, column=col_dest_fecha).value = dato['serie']
                ventas_ws.cell(row=fila_escritura, column=col_dest_nombre).value = dato['nombre']
                
                numero_limpio = ""
                try:
                    numero_limpio = int(float(dato['numero'])) if dato['numero'] else ""
                except ValueError:
                    numero_limpio = dato['numero']
                
                if dato['tipo'] == 'FACTURA':
                    ventas_ws.cell(row=fila_escritura, column=col_dest_n_fact).value = numero_limpio
                    ventas_ws.cell(row=fila_escritura, column=col_dest_monto_f).value = dato['total']
                    cont_facturas += 1
                
                elif dato['tipo'] == 'NC':
                    ventas_ws.cell(row=fila_escritura, column=col_dest_n_nc).value = numero_limpio
                    ventas_ws.cell(row=fila_escritura, column=col_dest_monto_nc).value = dato['total']
                    cont_nc += 1
        else:
            print(f"\n⚠️ NO SE ENCONTRÓ el marcador final para el día {dia}. Registros ignorados.")

    total_registros = cont_facturas + cont_nc
    if total_registros > 0:
        ventas_wb.save(archivo_destino)
        print(f"\n✅ ¡ÉXITO! Archivo '{archivo_destino}' actualizado con {cont_facturas} Facturas y {cont_nc} NC.")
        mover_procesados(archivos_originales)


# ==========================================
# 3. MÓDULO: ARCHIVOS 'CIERRE TOTAL' -> LOCALES DETALLE Y CONTROL DE EFECTIVO
# ==========================================

def obtener_diccionario_metodos(ventas_ws, day):
    total_indices = []
    for fila in range(3, ventas_ws.max_row + 1):
        val_col1 = ventas_ws.cell(row=fila, column=1).value
        val_col2 = ventas_ws.cell(row=fila, column=2).value
        val = val_col1 or val_col2
        if val and "TOTAL" in str(val).strip().upper():
            total_indices.append(fila)

    if not total_indices:
        print("⚠️ No se encontraron filas con 'TOTAL' en columna 1 o 2.")
        return {}

    inicio = 3 if day == 1 else total_indices[day - 2] + 1
    fin = total_indices[day - 1] 

    filas_dia = list(range(inicio, fin + 1))
    nombres_metodos = []
    
    for fila in filas_dia:
        val_col2 = ventas_ws.cell(row=fila, column=2).value
        val_col1 = ventas_ws.cell(row=fila, column=1).value
        val = val_col2 or val_col1
        if val and str(val).strip():
            nombres_metodos.append(str(val).strip())

    metodo_pago_map = {}
    for idx, metodo in enumerate(nombres_metodos):
        metodo_pago_map[normalizar(metodo)] = (metodo.replace(" ", ""), filas_dia[idx])

    return metodo_pago_map

def detectar_cajas(cierre_ws):
    def try_row(row):
        found = []
        for col in range(3, cierre_ws.max_column + 1):
            cell = cierre_ws.cell(row=row, column=col).value
            if cell is None:
                continue
            raw = str(cell).strip()
            raw_up = raw.upper()

            if "RYV" in raw_up:
                m = re.search(r'(\d+)\s*$', raw)
                if m: found.append((col, str(int(m.group(1)))))
                continue

            if "-" in raw:
                left = raw.split("-", 1)[0].strip()
                m = re.search(r'(\d+)', left)
                if m: found.append((col, str(int(m.group(1)))))
                else:
                    m2 = re.search(r'(\d+)', raw)
                    if m2: found.append((col, str(int(m2.group(1)))))
                continue

            m = re.search(r'(\d+)', raw)
            if m:
                found.append((col, str(int(m.group(1)))))
                continue
        return found

    cajas = try_row(3)
    if not cajas:
        for r in (1, 2, 4):
            cajas = try_row(r)
            if cajas: break
    return cajas

def transfer_control_efectivo_integrado(cierre_files, ventas_wb):
    """Extrae Vendedores y Efectivo de CIERRE TOTAL y los manda a CONTROL DE EFECTIVO."""
    print("\n--- Procesando Pestaña: CONTROL DE EFECTIVO (Vendedores) ---")
    try:
        control_ws = ventas_wb["CONTROL DE EFECTIVO"]
    except Exception as e:
        print(f"❌ Error: No se encontró la hoja 'CONTROL DE EFECTIVO'. {e}")
        return

    verde_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    amarillo_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for local, day, mes_nombre, cierre_file in cierre_files:
        try:
            cierre_wb = openpyxl.load_workbook(cierre_file, data_only=True)
            cierre_ws = cierre_wb.active

            fila_vendedor = fila_caja = fila_efectivo = None
            for fila in range(1, cierre_ws.max_row + 1):
                val = str(cierre_ws.cell(row=fila, column=1).value or "").upper()
                if val == "VENDEDOR":
                    fila_vendedor = fila
                elif val.startswith("ARQUEO Z"):
                    fila_caja = fila
                elif val == "EFECTIVO":
                    fila_efectivo = fila

            if not fila_vendedor or not fila_caja or not fila_efectivo:
                print(f"⚠️ Faltan filas clave (Vendedor/Efectivo) en {cierre_file}. Saltando.")
                continue

            start_col = int((day - 0.75) * 4)
            
            for col in range(3, cierre_ws.max_column + 1):
                vendedor = cierre_ws.cell(row=fila_vendedor, column=col).value
                caja_raw = cierre_ws.cell(row=fila_caja, column=col).value
                efectivo = cierre_ws.cell(row=fila_efectivo, column=col).value

                if not vendedor and not caja_raw and not efectivo:
                    continue  

                caja_num = str(caja_raw).split("-")[0].strip() if caja_raw else ""
                
                # BUSCAR LA PRIMERA FILA VACÍA PARA NO SOBRESCRIBIR EL LOCAL ANTERIOR
                row_dest = 4
                while (control_ws.cell(row=row_dest, column=start_col).value is not None or 
                       control_ws.cell(row=row_dest, column=start_col + 1).value is not None or 
                       control_ws.cell(row=row_dest, column=start_col + 2).value is not None):
                    row_dest += 1

                control_ws.cell(row=row_dest, column=start_col).value = vendedor
                control_ws.cell(row=row_dest, column=start_col + 1).value = caja_num
                control_ws.cell(row=row_dest, column=start_col + 2).value = _convertir_num(efectivo)

                fill = verde_fill if str(local) == "905" else amarillo_fill
                for c in range(start_col, start_col + 3):
                    control_ws.cell(row=row_dest, column=c).fill = fill

            print(f"  ✓ Cajeros y Efectivo traspasados para Día {day} - Local {local}")

        except Exception as e:
            print(f"❌ Error en efectivo de {cierre_file}: {e}")

def procesar_cierres_totales():
    print("\n" + "="*50)
    print("▶ INICIANDO: TRASPASO DE 'CIERRES TOTALES' (Ventas y Cajeros)")
    print("="*50)

    cierre_files = []
    archivos_originales = []
    pattern = re.compile(r"CIERRE\s*TOTAL\s+(\d+)\s+(\d{2})-(\d{2})", re.IGNORECASE)

    for file in os.listdir():
        if file.lower().endswith((".xls", ".xlsx")) and not file.startswith("~"):
            match = pattern.search(file)
            if match:
                local, dia, mes = match.groups()
                mes_nombre = MESES_ES.get(mes)
                if mes_nombre:
                    cierre_files.append((local, int(dia), mes_nombre, file))
                    archivos_originales.append(file)

    if not cierre_files:
        print("⚠️ No se encontraron archivos 'CIERRE TOTAL' nuevos.")
        return

    # Ordenar los archivos: Por Día, y luego por Local (905 antes que 1887)
    cierre_files.sort(key=lambda x: (x[1], x[0]))

    mes_nombre_dest = cierre_files[0][2]
    ano_actual = datetime.now().year
    archivo_destino = f"{mes_nombre_dest} {ano_actual}.xlsx"

    try:
        ventas_wb = openpyxl.load_workbook(archivo_destino)
        ventas_ws = ventas_wb["LOCALES DETALLE"]
        print(f"✓ Archivo maestro cargado: {archivo_destino}")
    except Exception as e:
        print(f"❌ Error al cargar maestro para cierres: {e}")
        return

    dest_cols_map = defaultdict(list)
    for col in range(3, ventas_ws.max_column + 1):
        header = ventas_ws.cell(row=2, column=col).value
        if header is None:
            continue
        m = re.search(r'(\d+)', str(header).strip())
        if m:
            dest_cols_map[str(int(m.group(1)))].append(col)

    print("\n--- Procesando Pestaña: LOCALES DETALLE ---")
    for local, day, mes_nombre, cierre_file in cierre_files:
        try:
            cierre_wb = openpyxl.load_workbook(cierre_file, data_only=True)
            cierre_ws = cierre_wb.active

            metodo_pago_map = obtener_diccionario_metodos(ventas_ws, day)
            if not metodo_pago_map:
                continue

            cajas = detectar_cajas(cierre_ws)
            cierre_data = {}
            
            for metodo_norm, (metodo_destino, target_row) in metodo_pago_map.items():
                valores = []
                encontrado = False
                for row in range(1, cierre_ws.max_row + 1):
                    metodo_cell = str(cierre_ws.cell(row=row, column=1).value or "").strip()
                    if normalizar(metodo_cell) == metodo_norm:
                        encontrado = True
                        for col, caja_num in cajas:
                            valores.append((caja_num, cierre_ws.cell(row=row, column=col).value))
                        break
                if not encontrado:
                    for col, caja_num in cajas:
                        valores.append((caja_num, None))
                cierre_data[metodo_destino] = (target_row, valores)

            for metodo, (target_row, valores) in cierre_data.items():
                if target_row is None:
                    continue
                counters = defaultdict(int)
                for caja_num, valor in valores:
                    dest_list = dest_cols_map.get(str(caja_num))
                    if not dest_list:
                        continue
                    idx = counters[str(caja_num)]
                    if idx < len(dest_list):
                        ventas_ws.cell(row=target_row, column=dest_list[idx]).value = _convertir_num(valor)
                        counters[str(caja_num)] += 1
            
            print(f"  ✓ Ventas por métodos traspasadas para Día {day} - Local {local}")

        except Exception as e:
            print(f"❌ Error procesando ventas de {cierre_file}: {e}")

    # Ahora sí, aquí se inyecta la lógica de Control de Efectivo
    transfer_control_efectivo_integrado(cierre_files, ventas_wb)

    try:
        ventas_wb.save(archivo_destino)
        print(f"\n✅ ¡ÉXITO! Cierres y Efectivo guardados en '{archivo_destino}'.")
        mover_procesados(archivos_originales)
    except Exception as e:
        print(f"\n❌ Error Crítico al guardar: {e}. Asegúrate de que el Excel esté cerrado.")


# ==========================================
# 4. ORQUESTADOR MAESTRO
# ==========================================

def main():
    print("****************************************************************")
    print("          🤖 AUTOMATIZADOR MAESTRO DE CONTABILIDAD 🤖           ")
    print("****************************************************************")
    print("El sistema procesará automáticamente:")
    print(" 1. Facturas y NC -> Hoja 'FACTURA A COBRAR' (Facturas primero)")
    print(" 2. Archivos 'Cierre Total' -> Hojas 'LOCALES DETALLE' y 'CONTROL DE EFECTIVO'")
    print("----------------------------------------------------------------")
    
    procesar_facturas_y_nc()
    procesar_cierres_totales()
        
    print("\n****************************************************************")
    print("                 ✅ PROCESO TOTAL FINALIZADO ✅                  ")
    print("****************************************************************")
    
    input("\nPresiona Enter para salir y cerrar esta ventana...")

if __name__ == "__main__":
    main()